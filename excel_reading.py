###This script will enable the reading of the clinical data from the excel files
###This section of the code was written by Aidan Riley
#IMPORTANT NOTE
#line 26 of the march clinical data excel file needs to be corrected so that the two numbers are spaced apart

from load import import_files
from openpyxl import load_workbook
import os
import copy

#functions for pairing row names with indicies
def col_name_dict(workbook):
    for row in workbook.iter_rows(min_row=1,max_row=1,values_only=True):
        tup = row
    dict = {tup[i]:i for i in range(len(tup))}
    return dict
def type_help(item):
    out = []
    for word in item.split():
        if ',' in word:
            word = word[:-1]
            out.append(int(word))
        elif word.isdigit():
            out.append(int(word))
    return out

#A class that stores the information from the Clinical Data Excel sheet for the March Data
class Patient:
    def __init__(self, data, month):
        if month == 'march':
            self.mz = None
            self.intensity = None
            self.type = None
            self.month = month
            #needed this line because bank_nums was imported as a string
            self.bank_nums = type_help(data[0])[0],type_help(data[0])[1]
            self.age = data[1]
            self.treatment = data[2]
            self.cycle = data[3]
            self.t_size = data[4]
            self.nodal_status = data[5]
            self.ER = data[6]
            self.PR = data[7]
            self.HER2_IHC = data[8]
            self.HER2_FISH = data[9]
            self.histo = data[10]
            self.clinical_response = data[11]
            self.pathologic_response = data[12]
        elif month == 'junepre':
            self.mz = None
            self.intensity = None
            self.type = None
            self.month = month
            self.bank_nums = (data[0],data[1])
            self.age = data[3]
            self.treatment = data[5]
            self.cycle = data[6]
            self.t_size = data[7]
            self.nodal_status = data[8]
            self.ER = data[9]
            self.PR = data[10]
            self.HER2_IHC = data[11]
            self.HER2_FISH = data[12]
            self.histo = data[13]
            self.clinical_response = data[14]
            self.pathologic_response = data[15]
        elif month == 'junepost':
            self.mz = None
            self.intensity = None
            self.type = None
            self.month = month
            self.bank_nums = (data[0],data[1])
            self.age = data[3]
            self.treatment = data[5]
            self.cycle = data[6]
            self.t_size = data[7]
            self.nodal_status = data[8]
            self.ER = data[9]
            self.PR = data[10]
            self.HER2_IHC = data[11]
            self.HER2_FISH = data[12]
            self.histo = data[13]
            #no clinical/pathological response in post-op june sheet
            #self.clinical_response = data[14]
            #self.pathologic_response = data[15]

    def __str__(self):
        return '%s %s %s \n' % (self.month, self.bank_nums, self.type)




#this will extract the clinical data from a specific excel file
#type can be equal to 'june' or 'march'
#output is a list of patient objects
def parse_file(file,ver):
    output = []
    #go through each row, and create a patient for the values in each column of that row
    for row in file.iter_rows(min_row=2, max_row=file.max_row, values_only=True):
        if row[0] is None:
            pass
        else:
            output.append(Patient(row,ver))
    return output
#this function will call parse_file on the entire folder that the data is stored in
def parse_folder(folder):
    file_list = os.listdir(folder)
    obj_lst = []
    for file in file_list:
        if file[0:6] == 'Normal':
            pass
        elif file[-11:-7] == 'june':
            ver = 'junepre'
            book_one = load_workbook('%s' % (folder+'\\'+file))['Pre-op FAC']
            obj_lst = obj_lst + parse_file(book_one,ver)

            ver = 'junepost'
            book_two = load_workbook('%s' % (folder+'\\'+file))['Post-op']
            obj_lst = obj_lst + parse_file(book_two,ver)
        else:
            ver = 'march'
            book = load_workbook('%s' % (folder + '\\' + file))['Sheet1']
            obj_lst = obj_lst + parse_file(book, ver)
    return obj_lst



#restructuring the object
#takes the output of parse folder as an argument
def rework(obj_lst):
    out_dict = {}
    #can't destructively copy, needed to use deepcopy() method
    for patient in obj_lst:
        key_1 = patient.bank_nums[0]
        key_2 = patient.bank_nums[1]
        patient_copy = copy.deepcopy(patient)
        patient.bank_nums = key_1
        patient_copy.bank_nums = key_2
        out_dict[key_1] = patient
        out_dict[key_2] = patient_copy
    return out_dict


#import the june and march low intensity xml files
#enter the path to the xml files here
june_xml_low = import_files(["low_mass_june03\\LowPrePostJune03\\RawXML"])
june_xml_high = import_files(["high_mass_june03\\HighPrePostJune03\\RawXML"])
march_xml_low = import_files(["low_mass_march03\\LowNorPrePost\\RawXML"])
march_xml_high = import_files(["high_mass_march03\\HighNorPrePost\\RawXML"])
#create one combined list of all the xml objects
combined_xml_lst = june_xml_low + june_xml_high + march_xml_low + march_xml_high

#enter the name of your clinical data here
y = parse_folder("clinical_data")
y = rework(y)

#convert the sample data type into a dictionary for lookup
def samples_to_dict(sample_lst):
    dict = {}
    for sample in sample_lst:
        dict[int(sample.name[2:])] = sample
    return dict




def link(xml,excel_lst):
    lookup = samples_to_dict(xml)
    for key in excel_lst:
        patient = excel_lst[key]
        #Retrieve the bank number identifier
        num = patient.bank_nums
        if num == 'N/A':
            pass
        else:
            # Access the corresponding xml data
            sample_dat = lookup[num]
            # add the xml data to the patient object
            patient.mz = sample_dat.mz
            patient.intensity = sample_dat.intensity
            patient.type = sample_dat.type_f

link(combined_xml_lst,y)

#removing the dictionary format, optional
def undo_dictionary(dict):
    out = []
    for key in dict:
        out.append(dict[key])
    return out
y = undo_dictionary(y)


